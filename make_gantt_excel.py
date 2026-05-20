"""
Generate an editable Gantt chart in Excel (.xlsx).
Fully editable: cells, colors, text, column widths — everything.
"""

from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# ── Colors ────────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FF0000")
GREEN_FILL = PatternFill("solid", fgColor="00B050")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HDR_FILL   = PatternFill("solid", fgColor="1F497D")   # dark blue header
ACT_FILL   = PatternFill("solid", fgColor="FFFFFF")

WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
BLACK_BOLD = Font(bold=True, color="000000", size=11)
BLACK_NORM = Font(bold=False, color="000000", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = thin_border()


# ── Column widths ─────────────────────────────────────────────
ws.column_dimensions["A"].width = 22   # Activities
for col in ["B","C","D","E","F","G"]:
    ws.column_dimensions[col].width = 16


# ── Row 1 — Title ─────────────────────────────────────────────
ws.row_dimensions[1].height = 28
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "FACEE PROJECT — GANTT CHART"
c.font  = Font(bold=True, color="FFFFFF", size=14)
c.fill  = PatternFill("solid", fgColor="1F497D")
c.alignment = CENTER
c.border = BORDER


# ── Row 2 — Sub-title ─────────────────────────────────────────
ws.row_dimensions[2].height = 20
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "Red = Planned Schedule     Green = Actual Progress"
c.font  = Font(italic=True, color="404040", size=10)
c.fill  = PatternFill("solid", fgColor="D6DCE4")
c.alignment = CENTER
c.border = BORDER


# ── Row 3 — Column Headers ────────────────────────────────────
ws.row_dimensions[3].height = 40
headers = ["Activities",
           "28 Jan – 7 Feb",
           "8 Feb – 19 Mar",
           "20 Mar – 2 May",
           "3 May – 6 Jun",
           "7 Jun – 4 July",
           "5 Jul – 10 Jul"]
for ci, hdr in enumerate(headers, start=1):
    c = ws.cell(row=3, column=ci, value=hdr)
    c.font      = WHITE_FONT
    c.fill      = HDR_FILL
    c.alignment = CENTER
    c.border    = BORDER


# ── Activity data ─────────────────────────────────────────────
#  Each activity: (name, planned_cols, actual_cols)
#  cols are 1-based offsets into period columns B-G (so 1=B,2=C,...6=G)
ACTIVITIES = [
    ("Feasibility Study",       [1,2],       [1,2]),
    ("Literature Survey",       [1,2,4,5],   [1,2]),
    ("Interface Design",        [2,3],       [2]),
    ("Database Development",    [2,3,4,5],   [2,3]),
    ("Module Integration",      [4,5],       []),
    ("Testing and Deployment",  [5,6],       []),
    ("Documentation",           [1,2,4,5,6], [1,2]),
]

# Starting at Excel row 4; each activity = 2 rows (planned + actual)
START_ROW = 4

for ai, (name, planned, actual) in enumerate(ACTIVITIES):
    top_row = START_ROW + ai * 2      # planned row
    bot_row = START_ROW + ai * 2 + 1  # actual row

    ws.row_dimensions[top_row].height = 22
    ws.row_dimensions[bot_row].height = 22

    # Merge activity name over 2 rows
    merge_range = f"A{top_row}:A{bot_row}"
    ws.merge_cells(merge_range)
    c = ws.cell(row=top_row, column=1, value=name)
    c.font      = BLACK_BOLD
    c.fill      = WHITE_FILL
    c.alignment = LEFT
    c.border    = BORDER

    # Period cells — planned (top row)
    for ci in range(1, 7):
        cell = ws.cell(row=top_row, column=ci + 1)
        cell.fill   = RED_FILL if ci in planned else WHITE_FILL
        cell.border = BORDER

    # Period cells — actual (bottom row)
    for ci in range(1, 7):
        cell = ws.cell(row=bot_row, column=ci + 1)
        cell.fill   = GREEN_FILL if ci in actual else WHITE_FILL
        cell.border = BORDER

    # Re-apply border to merged A cell bottom
    ws.cell(row=bot_row, column=1).border = BORDER


# ── Legend rows ───────────────────────────────────────────────
leg_row = START_ROW + len(ACTIVITIES) * 2 + 1
ws.row_dimensions[leg_row].height = 22

ws.merge_cells(f"A{leg_row}:A{leg_row}")
c = ws.cell(row=leg_row, column=1, value="Legend:")
c.font = BLACK_BOLD; c.alignment = LEFT; c.border = BORDER
c.fill = PatternFill("solid", fgColor="F2F2F2")

ws.cell(row=leg_row, column=2).fill   = RED_FILL
ws.cell(row=leg_row, column=2).border = BORDER
ws.merge_cells(f"C{leg_row}:D{leg_row}")
c = ws.cell(row=leg_row, column=3, value="Planned Schedule")
c.font = BLACK_NORM; c.alignment = LEFT; c.border = BORDER
c.fill = PatternFill("solid", fgColor="F2F2F2")
ws.cell(row=leg_row, column=4).border = BORDER

ws.cell(row=leg_row, column=5).fill   = GREEN_FILL
ws.cell(row=leg_row, column=5).border = BORDER
ws.merge_cells(f"F{leg_row}:G{leg_row}")
c = ws.cell(row=leg_row, column=6, value="Actual Progress")
c.font = BLACK_NORM; c.alignment = LEFT; c.border = BORDER
c.fill = PatternFill("solid", fgColor="F2F2F2")
ws.cell(row=leg_row, column=7).border = BORDER


# ── Freeze panes at B4 (keep header + activities col visible) ─
ws.freeze_panes = "B4"

# ── Print settings ────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.print_title_rows       = "1:3"


out = "/home/user/FACEE/FACEE_Gantt_Chart.xlsx"
wb.save(out)
print(f"Saved → {out}")
